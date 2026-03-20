from openpyxl import load_workbook
import pandas as pd
import numpy as np

MONTH_MAP = {
    "102025": "2025-10",
    "112025": "2025-11",
    "122025": "2025-12",
    "012026": "2026-01",
    "022026": "2026-02",
    "032026": "2026-03",
}

MONTHLY_EVENTS = {
    "2025-10": {"event": "none",                "score": 0, "holiday_flag": 0},
    "2025-11": {"event": "none",                "score": 0, "holiday_flag": 0},
    "2025-12": {"event": "pre_new_year",         "score": 3, "holiday_flag": 1},
    "2026-01": {"event": "new_year_block",       "score": 3, "holiday_flag": 1},
    "2026-02": {"event": "valentines_defenders", "score": 2, "holiday_flag": 1},
    "2026-03": {"event": "womens_day",           "score": 3, "holiday_flag": 1},
}

HOLIDAY_DAYS_PER_MONTH = {
    "2025-10": 0, "2025-11": 4, "2025-12": 7,
    "2026-01": 8, "2026-02": 2, "2026-03": 1,
}

SEASON_MAP = {
    "2025-10": "early_winter", "2025-11": "winter",
    "2025-12": "winter_peak",  "2026-01": "post_new_year",
    "2026-02": "late_winter",  "2026-03": "spring_transition",
}

WINTER_SENSITIVITY = {
    "jacket": 3, "coat": 3, "sweater": 3, "cardigan": 2,
    "sweatshirt": 2, "vest": 2, "gloves": 3, "hat": 3,
    "long-sleeve polo shirt": 2, "long sleeve shirt": 1,
    "trousers": 1, "sports trousers": 1, "T-shirt": -1,
    "polo shirt": 0, "shirt": 0, "blouse": 0, "cap": 1,
    "bag": 0, "belt": 0, "jeans": 0, "scarf": 2,
}

ACCESSORY_CATS = {"bag", "belt", "cap", "hat", "gloves", "scarf"}

# Satıcı olmayan satır başlangıçları — kesin filtre
NON_MANAGER_PREFIXES = (
    "NS", "EU ", "EU S", "Size", "IT ", "Color",
    "Women", "Men", "Administrator", "Manager",
    "Nomenclature", "Characteristic",
)

def is_manager_name(val, qty, row1_is_none):
    """
    Gerçek satıcı ismi tespiti.
    Kriter: satır 1 None, qty sayı, ve isim beden/renk/ürün formatında değil.
    """
    if not val or not row1_is_none or not qty or not isinstance(qty, (int, float)):
        return False
    val = str(val).strip()

    # Kesin filtreler — bu prefix'lerle başlayan satırlar asla satıcı değil
    for prefix in NON_MANAGER_PREFIXES:
        if val.startswith(prefix):
            return False

    # Bilinen özel isimler
    if val == "Total":
        return False

    # Administrator özel durum (satıcı gibi davranır)
    if val == "Administrator":
        return True

    # Gerçek satıcı formatı: en az 2 büyük harf kelimesi olan tam isim
    words = val.split()
    upper_words = [w for w in words if w.isupper() and len(w) > 1]
    return len(upper_words) >= 2

def parse_sheet(ws, month_label):
    rows = list(ws.iter_rows(values_only=True))

    # İki geçiş: önce gerçek satıcıları tespit et
    known_managers = set()
    for row in rows:
        val0 = str(row[0]).strip() if row[0] else ""
        qty  = row[6]
        if is_manager_name(val0, qty, row[1] is None):
            known_managers.add(val0)

    records = []
    current_manager = None
    for i, row in enumerate(rows):
        val0 = str(row[0]).strip() if row[0] else ""
        qty  = row[6]
        rev  = row[8]
        full = row[9]
        unit = str(row[3]).strip() if row[3] else ""

        if val0 in known_managers:
            current_manager = val0
            continue

        if unit == "pcs" and qty and rev and isinstance(qty, (int, float)) and qty > 0:
            article = ""
            if i + 1 < len(rows):
                nxt = str(rows[i+1][0]).strip() if rows[i+1][0] else ""
                if nxt.startswith("NS"):
                    article = nxt

            discount_pct = round((1 - rev / full) * 100, 1) if full and full != 0 else 0.0
            product_clean = val0.strip("' ").strip()
            gender   = "Women" if "Women's" in product_clean else ("Men" if "Men's" in product_clean else "Unisex")
            category = product_clean.replace("Women's ", "").replace("Men's ", "").strip()

            records.append({
                "month": month_label, "manager": current_manager,
                "product": product_clean, "gender": gender,
                "category": category, "article": article,
                "quantity": int(qty), "revenue_rub": float(rev),
                "full_price_rub": float(full), "discount_pct": discount_pct,
            })
    return pd.DataFrame(records)

def load_all_months(path):
    wb = load_workbook(path, read_only=True)
    all_dfs = []
    for sheet_name in wb.sheetnames:
        clean = sheet_name.strip()
        if clean not in MONTH_MAP:
            print(f"⚠  Sheet '{clean}' MONTH_MAP'te yok, atlanıyor.")
            continue
        month_label = MONTH_MAP[clean]
        ws = wb[sheet_name]
        df = parse_sheet(ws, month_label)
        print(f"✓  {clean} → {month_label} | {len(df)} işlem | adet: {df['quantity'].sum()} | gelir: ₽{df['revenue_rub'].sum():,.0f}")
        all_dfs.append(df)
    combined = pd.concat(all_dfs, ignore_index=True)
    combined["period"] = pd.to_datetime(combined["month"])
    return combined.sort_values("period").reset_index(drop=True)

def add_monthly_calendar_features(df):
    df["holiday_flag"]       = df["month"].map(lambda m: MONTHLY_EVENTS.get(m, {}).get("holiday_flag", 0))
    df["holiday_importance"] = df["month"].map(lambda m: MONTHLY_EVENTS.get(m, {}).get("score", 0))
    df["holiday_name"]       = df["month"].map(lambda m: MONTHLY_EVENTS.get(m, {}).get("event", "none"))
    df["holiday_days_count"] = df["month"].map(lambda m: HOLIDAY_DAYS_PER_MONTH.get(m, 0))
    df["season"]             = df["month"].map(lambda m: SEASON_MAP.get(m, "unknown"))
    months_sorted = sorted(df["month"].unique())
    df["month_index"] = df["month"].map({m: i for i, m in enumerate(months_sorted)})
    df["month_num"]   = df["period"].dt.month
    df["quarter"]     = df["period"].dt.quarter
    return df

def add_product_features(df):
    df["winter_sensitivity"] = df["category"].map(lambda c: WINTER_SENSITIVITY.get(c, 0))
    df["margin_proxy"]       = 1 - df["discount_pct"] / 100
    df["avg_ticket_rub"]     = df["revenue_rub"] / df["quantity"]
    df["is_premium"]         = (df["full_price_rub"] / df["quantity"] > 15000).astype(int)
    df["is_accessory"]       = df["category"].isin(ACCESSORY_CATS).astype(int)
    df["discount_band"]      = pd.cut(
        df["discount_pct"],
        bins=[-1, 0, 15, 35, 55, 100],
        labels=["full_price", "light", "moderate", "deep", "clearance"]
    )
    return df

def add_manager_features(df):
    mgr_monthly = (df.groupby(["month","manager"])["revenue_rub"].sum()
                   .reset_index().rename(columns={"revenue_rub":"mgr_monthly_revenue"}))
    df = df.merge(mgr_monthly, on=["month","manager"], how="left")
    total_monthly = df.groupby("month")["revenue_rub"].transform("sum")
    df["mgr_revenue_share"] = df["mgr_monthly_revenue"] / total_monthly
    mgr_avg_disc = (df.groupby(["month","manager"])["discount_pct"].mean()
                    .reset_index().rename(columns={"discount_pct":"mgr_avg_discount"}))
    df = df.merge(mgr_avg_disc, on=["month","manager"], how="left")
    return df

def add_lag_features(df):
    cat_monthly = (df.groupby(["month","category"])
                   .agg(cat_qty=("quantity","sum"), cat_rev=("revenue_rub","sum"))
                   .reset_index().sort_values("month"))
    cat_monthly["cat_rev_lag1"]   = cat_monthly.groupby("category")["cat_rev"].shift(1)
    cat_monthly["cat_rev_lag2"]   = cat_monthly.groupby("category")["cat_rev"].shift(2)
    cat_monthly["cat_rev_growth"] = cat_monthly["cat_rev"] / cat_monthly["cat_rev_lag1"].replace(0, np.nan) - 1
    df = df.merge(cat_monthly[["month","category","cat_rev_lag1","cat_rev_lag2","cat_rev_growth"]],
                  on=["month","category"], how="left")
    store_monthly = (df.groupby("month")["revenue_rub"].sum()
                     .reset_index().rename(columns={"revenue_rub":"store_total_rev"}).sort_values("month"))
    store_monthly["store_rev_lag1"] = store_monthly["store_total_rev"].shift(1)
    store_monthly["store_growth"]   = store_monthly["store_total_rev"] / store_monthly["store_rev_lag1"].replace(0, np.nan) - 1
    df = df.merge(store_monthly[["month","store_total_rev","store_rev_lag1","store_growth"]],
                  on="month", how="left")
    return df

def build_monthly_targets(df):
    cat_agg = (df.groupby(["month","category"])
               .agg(total_qty=("quantity","sum"), total_rev=("revenue_rub","sum"),
                    avg_discount=("discount_pct","mean"))
               .reset_index().sort_values(["category","month"]))
    cat_agg["target_next_rev"] = cat_agg.groupby("category")["total_rev"].shift(-1)
    cat_agg["target_next_qty"] = cat_agg.groupby("category")["total_qty"].shift(-1)
    return cat_agg

# ─── ÇALIŞTIR ───
print("=" * 60)
print("  Fashion Retail Store — Feature Engineering Pipeline v3")
print("=" * 60 + "\n")

df = load_all_months('/mnt/user-data/uploads/Sales.xlsx')
print(f"\nToplam: {len(df)} işlem, {df['month'].nunique()} ay\n")

df = add_monthly_calendar_features(df)
df = add_product_features(df)
df = add_manager_features(df)
df = add_lag_features(df)
targets = build_monthly_targets(df)

df.to_csv('/home/claude/transactions_features.csv', index=False)
targets.to_csv('/home/claude/monthly_targets.csv', index=False)

# ─── ÖZET RAPOR ───
print("\n" + "=" * 60)
print("  ÖZET RAPOR")
print("=" * 60)

print("\n📅 AYLIK ÖZET:")
monthly = df.groupby("month").agg(
    islem=("quantity","count"),
    adet=("quantity","sum"),
    gelir=("revenue_rub","sum"),
    ort_indirim=("discount_pct","mean"),
    satici=("manager","nunique"),
    kategori=("category","nunique"),
).reset_index()
print(monthly.to_string(index=False))

print("\n🏷️  KATEGORİ BAZLI (tüm aylar):")
cat = df.groupby("category").agg(
    adet=("quantity","sum"), gelir=("revenue_rub","sum"),
    ort_indirim=("discount_pct","mean"),
).sort_values("gelir", ascending=False)
print(cat.to_string())

print("\n👤 SATICI BAZLI (tüm aylar):")
mgr = df.groupby("manager").agg(
    adet=("quantity","sum"), gelir=("revenue_rub","sum"),
    ort_indirim=("discount_pct","mean"),
).sort_values("gelir", ascending=False)
print(mgr.to_string())

# Satıcı doğrulama — sayısal/beden ismi kalmış mı?
print("\n⚠  Kontrol — beklenmedik satıcı isimleri:")
suspicious = [m for m in df["manager"].dropna().unique()
              if any(x in str(m) for x in ["Size", "Color", "EU ", "IT "])]
if suspicious:
    for s in suspicious:
        print(f"  → {s}")
else:
    print("  Temiz, sorun yok.")

print(f"\n📊 Feature matrisi: {df.shape[0]} satır × {df.shape[1]} kolon")
print(f"📊 Hedef matrisi:   {targets.shape[0]} satır × {targets.shape[1]} kolon")
